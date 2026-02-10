"""
Joom 형식 엑셀 출력기
"""
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

import sys
sys.path.append(str(Path(__file__).parent.parent))

from config import ExcelConfig, OUTPUT_DIR
from utils.logger import get_logger
from utils.color_mapping import ColorMapper


class JoomExcelExporter:
    """Joom 플랫폼 형식 엑셀 출력기"""

    def __init__(self, site_type: str = "musinsa"):
        """
        초기화

        Args:
            site_type: 사이트 유형 ("musinsa" 또는 "naver")
        """
        self.site_type = site_type
        self.config = ExcelConfig()
        self.logger = get_logger("ExcelExporter")
        self.color_mapper = ColorMapper()

        # 헤더 스타일
        self.header_font = Font(bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def get_headers(self) -> List[str]:
        """사이트 유형에 따른 헤더 목록 반환"""
        # 기본 헤더
        base_headers = [
            'Product_SKU',      # 부모 상품 ID
            'Variant_SKU',      # 옵션별 고유 ID
            'Product_Name',     # 상품명
            'Category',         # 카테고리 (신발>스니커즈>패션스니커즈화)
            'Brand',            # 브랜드
            'Color',            # 색상 (영문)
            'Size',             # 사이즈
            'Option1_Value',    # 첫 번째 옵션 값
            'Option2_Value',    # 두 번째 옵션 값
            'Option3_Value',    # 세 번째 옵션 값
            'Price',            # 정가
            'Sale_Price',       # 판매가
            'Additional_Price', # 추가금액
            'Total_Price',      # 총 가격
            'Stock',            # 재고
            'SoldOut',          # 품절여부
        ]

        # 사이트별 추가 헤더
        if self.site_type == "musinsa":
            extra_headers = [
                'StyleNo',          # 품번
                'Gender',           # 성별
                'Season',           # 시즌
                'ViewCount',        # 조회수
                'SalesCount',       # 누적판매
                'ReviewCount',      # 리뷰 개수
                'Rating',           # 평점
                'Description',      # 상세 설명 텍스트
            ]
        else:  # naver
            extra_headers = [
                # 상품정보 (viewAttributes)
                'ProductStatus',    # 상품상태
                'Manufacturer',     # 제조사
                'ModelName',        # 모델명
                'ManufactureDate',  # 제조일자
                'Origin',           # 원산지
                # 상품정보제공고시 (productInfoProvidedNoticeView)
                'Material',         # 제품소재
                'ColorInfo',        # 색상 (상품정보)
                'MadeBy',           # 제조자(사)
                'MadeIn',           # 제조국
                # 상세속성 (detailAttributes)
                'AnkleHeight',      # 발목높이
                'HeelHeight',       # 굽높이
                'MainMaterial',     # 주요소재
                'Function',         # 부가기능
                'Sole',             # 솔
                'ReviewCount',      # 리뷰 개수
                'Rating',           # 평점
                'Description',      # 상세 설명 텍스트
            ]

        # 공통 마지막 헤더
        common_end = [
            'Image_URL',        # 대표 이미지
            'Extra_Images',     # 추가 이미지 - 갤러리 (콤마 구분)
            'Detail_Images',    # 상세 설명 이미지 (콤마 구분)
            'URL',              # 상품 URL
            'Category_URL',     # 카테고리 URL
        ]

        return base_headers + extra_headers + common_end

    def export(self, products: List[Dict], output_path: Path = None, category_url: str = None) -> Path:
        """
        상품 데이터를 Joom 형식 엑셀로 출력

        Args:
            products: 상품 데이터 리스트
            output_path: 출력 파일 경로
            category_url: 수집한 카테고리 URL

        Returns:
            생성된 엑셀 파일 경로
        """
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = OUTPUT_DIR / f"{self.site_type}_products_{timestamp}.xlsx"

        self.logger.info(f"엑셀 파일 생성 시작: {output_path}")

        # 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.config.SHEET_NAME

        # 1행: 헤더
        headers = self.get_headers()
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border

        # 2행부터: 데이터
        row = 2
        total_options = 0

        for product in products:
            # ProductInfo 객체 또는 dict 모두 지원
            if hasattr(product, 'options'):
                options = product.options
            else:
                options = product.get('options', [])

            if options:
                # 옵션이 있는 경우: 옵션 개수만큼 행 생성
                for opt in options:
                    row_data = self._create_row_data(product, opt, category_url)
                    self._write_row(ws, row, row_data, headers)
                    row += 1
                    total_options += 1
            else:
                # 옵션이 없는 경우: 1개 행 생성
                row_data = self._create_row_data(product, None, category_url)
                self._write_row(ws, row, row_data, headers)
                row += 1

        # 열 너비 조정
        self._adjust_column_widths(ws, headers)

        # 저장
        wb.save(output_path)

        self.logger.info(f"엑셀 파일 생성 완료: {output_path}")
        self.logger.info(f"  총 상품: {len(products)}개")
        self.logger.info(f"  총 행 (옵션 확장): {row - 2}개")

        return output_path

    def _create_row_data(self, product, option, category_url: str = None) -> Dict:
        """행 데이터 생성 (ProductInfo 객체 또는 dict 모두 지원)"""
        # ProductInfo 객체 또는 dict 모두 지원
        if hasattr(product, 'product_id'):
            product_id = product.product_id
            extra_info = product.extra_info
            product_name = product.product_name
            brand = product.brand
            price = product.price
            sale_price = product.sale_price
            image_url = product.image_url
            url = product.url
        else:
            product_id = product.get('product_id', '')
            extra_info = product.get('extra_info', {})
            product_name = product.get('product_name', '')
            brand = product.get('brand', '')
            price = product.get('price', 0)
            sale_price = product.get('sale_price', 0)
            image_url = product.get('image_url', '')
            url = product.get('url', '')

        # 추가 이미지 (콤마로 구분)
        extra_images = extra_info.get('extra_images', [])
        extra_images_str = ', '.join(extra_images) if extra_images else ''

        # 상세 설명 이미지 (콤마로 구분)
        detail_images = extra_info.get('detail_images', [])
        detail_images_str = ', '.join(detail_images) if detail_images else ''

        # 기본 데이터
        data = {
            'Product_SKU': product_id,
            'Product_Name': product_name,
            'Category': extra_info.get('category', ''),
            'Brand': brand,
            'Price': price,
            'Sale_Price': sale_price,
            'Image_URL': image_url,
            'Extra_Images': extra_images_str,
            'Detail_Images': detail_images_str,
            'URL': url,
            'Category_URL': category_url or '',
        }

        # 옵션 데이터 (ProductOption 객체 또는 dict 모두 지원)
        if option:
            if hasattr(option, 'color'):
                color = option.color
                size = option.size
                additional_price = option.additional_price
                sold_out = option.sold_out
                stock = option.stock if option.stock else self.config.DEFAULT_STOCK
                option_data = option.option_data if hasattr(option, 'option_data') else {}
            else:
                color = option.get('color', '')
                size = option.get('size', '')
                additional_price = option.get('additional_price', 0)
                sold_out = option.get('sold_out', False)
                stock = option.get('stock', self.config.DEFAULT_STOCK)
                option_data = option.get('option_data', {})

            # 색상 영문 변환
            color_eng = self.color_mapper.to_english(color) if color else ''

            # Variant SKU 생성
            variant_parts = [product_id]
            if color:
                variant_parts.append(color[:3].upper())
            if size:
                variant_parts.append(str(size).replace(' ', ''))
            variant_sku = '-'.join(variant_parts)

            # option_data를 순서대로 Option1/2/3 Value에 배분
            opt_items = list(option_data.items()) if option_data else []
            opt1_value = opt_items[0][1] if len(opt_items) > 0 else ''
            opt2_value = opt_items[1][1] if len(opt_items) > 1 else ''
            opt3_value = opt_items[2][1] if len(opt_items) > 2 else ''

            data.update({
                'Variant_SKU': variant_sku,
                'Color': color_eng,
                'Size': size,
                'Option1_Value': opt1_value,
                'Option2_Value': opt2_value,
                'Option3_Value': opt3_value,
                'Additional_Price': additional_price,
                'Total_Price': data['Sale_Price'] + additional_price,
                'Stock': 0 if sold_out else stock,
                'SoldOut': 'Y' if sold_out else 'N',
            })

            # 옵션별 이미지 URL (다이소 등 옵션마다 이미지가 다른 경우)
            option_image = ''
            if hasattr(option, 'image_url'):
                option_image = option.image_url
            else:
                option_image = option.get('image_url', '')
            if option_image:
                data['Image_URL'] = option_image

            # 옵션별 추가 이미지 (갤러리)
            option_extra_images = []
            if hasattr(option, 'extra_images'):
                option_extra_images = option.extra_images
            else:
                option_extra_images = option.get('extra_images', [])
            if option_extra_images:
                data['Extra_Images'] = ', '.join(option_extra_images)
        else:
            # 옵션 없는 경우
            data.update({
                'Variant_SKU': product_id,
                'Color': '',
                'Size': '',
                'Option1_Value': '',
                'Option2_Value': '',
                'Option3_Value': '',
                'Additional_Price': 0,
                'Total_Price': data['Sale_Price'],
                'Stock': self.config.DEFAULT_STOCK,
                'SoldOut': 'N',
            })

        # 사이트별 추가 데이터
        if self.site_type == "musinsa":
            data.update({
                'StyleNo': extra_info.get('style_no', ''),
                'Gender': extra_info.get('gender', ''),
                'Season': extra_info.get('season', ''),
                'ViewCount': extra_info.get('view_count', 0),
                'SalesCount': extra_info.get('sales_count', 0),
                'ReviewCount': extra_info.get('review_count', 0),
                'Rating': extra_info.get('rating', 0),
                'Description': extra_info.get('description', ''),
            })
        else:  # naver
            data.update({
                # 상품정보 (viewAttributes)
                'ProductStatus': extra_info.get('product_status', ''),
                'Manufacturer': extra_info.get('manufacturer', ''),
                'ModelName': extra_info.get('model_name', ''),
                'ManufactureDate': extra_info.get('manufacture_date', ''),
                'Origin': extra_info.get('origin', ''),
                # 상품정보제공고시 (productInfoProvidedNoticeView)
                'Material': extra_info.get('material', ''),
                'ColorInfo': extra_info.get('color_info', ''),
                'MadeBy': extra_info.get('made_by', ''),
                'MadeIn': extra_info.get('made_in', ''),
                # 상세속성 (detailAttributes)
                'AnkleHeight': extra_info.get('ankle_height', ''),
                'HeelHeight': extra_info.get('heel_height_detail', ''),
                'MainMaterial': extra_info.get('main_material', ''),
                'Function': extra_info.get('function', ''),
                'Sole': extra_info.get('sole', ''),
                'ReviewCount': extra_info.get('review_count', 0),
                'Rating': extra_info.get('rating', 0),
                'Description': extra_info.get('description', ''),
            })

        return data

    def _write_row(self, ws, row: int, data: Dict, headers: List[str]):
        """행 쓰기"""
        for col, header in enumerate(headers, 1):
            value = data.get(header, '')
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = self.thin_border

            # 숫자 컬럼 우측 정렬
            if header in ['Price', 'Sale_Price', 'Additional_Price', 'Total_Price', 'Stock', 'ViewCount', 'SalesCount', 'ReviewCount', 'Rating']:
                cell.alignment = Alignment(horizontal='right')

    def _adjust_column_widths(self, ws, headers: List[str]):
        """열 너비 조정"""
        width_map = {
            'Product_SKU': 15,
            'Variant_SKU': 20,
            'Product_Name': 45,
            'Category': 40,
            'Brand': 15,
            'Color': 12,
            'Size': 10,
            'Option1_Value': 20,
            'Option2_Value': 20,
            'Option3_Value': 20,
            'Price': 12,
            'Sale_Price': 12,
            'Additional_Price': 15,
            'Total_Price': 12,
            'Stock': 8,
            'SoldOut': 8,
            'StyleNo': 15,
            'Gender': 12,
            'Season': 12,
            'ViewCount': 12,
            'SalesCount': 12,
            'ProductStatus': 12,
            'Origin': 15,
            'Manufacturer': 15,
            'ModelName': 20,
            'ManufactureDate': 14,
            'Material': 40,
            'ColorInfo': 30,
            'MadeBy': 15,
            'MadeIn': 12,
            'AnkleHeight': 12,
            'HeelHeight': 12,
            'MainMaterial': 15,
            'Function': 12,
            'Sole': 10,
            'ReviewCount': 12,
            'Rating': 10,
            'Description': 80,
            'Image_URL': 50,
            'Extra_Images': 100,
            'URL': 50,
            'Category_URL': 60,
        }

        for col, header in enumerate(headers, 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            ws.column_dimensions[col_letter].width = width_map.get(header, 15)


def export_to_excel(products: List[Dict], site_type: str, output_path: Path = None, category_url: str = None) -> Path:
    """
    상품 데이터를 엑셀로 출력하는 헬퍼 함수

    Args:
        products: 상품 데이터 리스트
        site_type: 사이트 유형 ("musinsa" 또는 "naver")
        output_path: 출력 파일 경로
        category_url: 수집한 카테고리 URL

    Returns:
        생성된 엑셀 파일 경로
    """
    exporter = JoomExcelExporter(site_type)
    return exporter.export(products, output_path, category_url)
